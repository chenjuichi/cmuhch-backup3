import os
import datetime
import pathlib
import csv

# ------------------------------------------------------------------

import openpyxl
from openpyxl import Workbook, load_workbook
# 載入 Font 和 PatternFill 模組
from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
from database.tables import InTag, Grid, Reagent, Session

from flask import Blueprint, jsonify, request

excelTable = Blueprint('excelTable', __name__)

# ------------------------------------------------------------------


def modify_InTags_grid(_id, _station, _layout, _pos, _reagID):
    return_gridID = 0  # 相同儲位

    s = Session()
    target_grid = s.query(Grid).filter_by(
        station=_station, layout=_layout, pos=_pos).first()

    if not target_grid:
        # new grid, 建立新的儲位
        new_grid = Grid(station=_station, layout=_layout, pos=_pos)
        s.add(new_grid)
        s.flush()
        current_reagent = s.query(Reagent).filter_by(reag_id=_reagID).first()
        current_reagent.grid_id = new_grid.id
        s.commit()
        return_gridID = new_grid.id
    elif not (target_grid.id == _id):  # target grid不等於既有的儲位, 就是不同儲位
        reagent_count = s.query(Reagent).filter_by(grid_id=target_grid.id).count()

        if reagent_count >= 1:  # 已經放其他試劑, 同一儲位不能放不同試劑
            print("hello, another same records...")
            return_gridID = -1  # 同一儲位, 不能放相同試劑
        else:
            return_gridID = target_grid.id  # 空儲位, 沒有放其他試劑

    s.close()

    return return_gridID


# ------------------------------------------------------------------


@excelTable.route("/exportToCSVForStockInOut", methods=['POST'])
def export_to_csv_for_stock_in_out():
    print("exportToCSVForStockInOut....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    print("b data: ", _blocks, _count)

    return_value = True  # true: write into csv成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        for obj in _blocks:

          print("excel print: ", obj)

          #if (obj['isIn']):                          # 2023-12-7 modify
          if (obj['stockInTag_rePrint']=='入庫'):     # 2023-12-7 modify
              obj['stockInTag_Employer'] = '入庫人員: ' + \
                  obj['stockInTag_Employer']
              obj['stockInTag_Date'] = '入庫日期: ' + \
                  obj['stockInTag_Date']
          else:
              obj['stockInTag_Employer'] = '領料人員: ' + \
                  obj['stockInTag_Employer']
              obj['stockInTag_Date'] = '領料日期: ' + \
                  obj['stockInTag_Date']

          obj['stockInTag_cnt'] = 1   # 2023-08-25 add

        print("a data: ", _blocks)

        for obj in _blocks:
            del obj['isIn']                 # 2022-12-23 remove 'isIn' key
            del obj['stockInTag_rePrint']   # 2023-08-10 remove 'stockInTag_rePrint' key
            del obj['stockInTag_reagName']  # 2023-08-10 remove 'stockInTag_reagName' key
        print("aa excel print: ", _blocks)

        # 查看當前工作目錄
        olddir = os.getcwd()
        print("當前工作目錄:%s" % olddir)

        myDrive = pathlib.Path.home().drive
        # print(myDrive)

        mypath = 'e:\\CMUHCH\\print.csv'
        #mypath = myDrive + '\\CMUHCH\\print.csv'
        myDir = 'e:\\CMUHCH\\'
        #myDir = myDrive + '\\CMUHCH\\'
        myFile = 'print.csv'

        if os.path.isdir(myDir):  #目錄存在
            if os.path.isfile(mypath):  #目錄內, 目標檔案存在
                os.chdir(myDir)         # 進入csv目錄
                # im.close()
                os.remove(myFile)       #刪除目標檔案
            else:
                #if not os.path.exists(myDir):
                #    os.makedirs(myDir)
                os.chdir(myDir)  # 進入csv目錄

            # 查看當前工作目錄
            curdir = os.getcwd()
            print("修改後工作目錄:%s" % curdir)

            # csvfile = open(myFile, 'a+')    # 使用 a+ 模式開啟檔案
            csvfile = open(myFile, 'w', newline='')
            try:
                fieldnames = ['name1', 'name2',
                              'stockInTag_reagID',
                              'stockInTag_batch',
                              'stockInTag_Date',
                              'stockInTag_Employer',
                              'stockInTag_reagTemp',
                              'stockInTag_alpha',
                              'stockInTag_cnt', ]  # 定義要寫入資料的鍵
                data = csv.DictWriter(
                    csvfile, fieldnames=fieldnames)  # 設定 data 為寫入資料
                data.writerows(_blocks)

            finally:
                csvfile.close()

            os.system("e:\\CMUHCH\\barcode.bat")

        else: #for test
          mypath = 'c:\\CMUHCH\\print.csv'
          myDir = 'c:\\CMUHCH\\'
          myFile = 'print.csv'

          os.chdir(myDir)         # 進入csv目錄

          ### 2023-12-7 modify the following block
          temp_file_name = myFile.replace(".csv", "_temp.csv")
          #if os.path.isfile(temp_file_name):
          os.getcwd()
          print("kk 當前工作目錄:%s" % os.getcwd())
          if os.path.exists(temp_file_name):
            os.remove(temp_file_name)       #刪除目標檔案

          if os.path.isfile(mypath):  #目錄內, 目標檔案存在

            try:
                os.remove(myFile)
            except PermissionError as e:
                print(f"無法刪除檔案: {e}")
                # 使用臨時檔案名稱
                myFile = temp_file_name

            #os.remove(myFile)       #刪除目標檔案
          ###
          csvfile = open(myFile, 'w', newline='')
          try:
              fieldnames = ['name1', 'name2',
                            'stockInTag_reagID',
                            'stockInTag_batch',
                            'stockInTag_Date',
                            'stockInTag_Employer',
                            'stockInTag_reagTemp',
                            'stockInTag_alpha',
                            'stockInTag_cnt', ]  # 定義要寫入資料的鍵
              data = csv.DictWriter(
                  csvfile, fieldnames=fieldnames)  # 設定 data 為寫入資料
              data.writerows(_blocks)

          finally:
              csvfile.close()

        os.chdir(olddir)  # 進入server工作目錄
        print("最後工作目錄:%s" % olddir)

    return jsonify({
        'status': return_value,
        'outputs': mypath,
    })


@excelTable.route("/exportToExcelForReq", methods=['POST'])
def export_to_Excel_for_Req():
    print("exportToExcelForReq....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    #current_file = '..\\領用記錄查詢_'+today + '.xlsx'
    current_file = 'e:\\CMUHCH\\領用記錄查詢_'+today + '.xlsx'
    #current_file0 = '領用記錄查詢_'+today + '.xlsx'

    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '領用記錄查詢-' + _name                # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            temp_array = []
            temp_array.append(obj['reqRecord_reagID'])
            temp_array.append(obj['reqRecord_reagName'])
            temp_array.append(obj['reqRecord_supplier'])
            temp_array.append(obj['reqRecord_stockInDate'])
            temp_array.append(obj['reqRecord_Date'])
            temp_array.append(obj['reqRecord_Employer'])
            # temp_str = obj['reqRecord_cnt'] + obj['reqRecord_unit']
            temp_array.append(obj['reqRecord_cnt'])

            ws.append(temp_array)
            # print("obj: ", temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True
            # dim = openpyxl.worksheet.dimensions.ColumnDimension(
            #    ws, index=column, bestFit=True, customWidth=True)
            # ws.column_dimensions[column] = dim

        wb.save(current_file)

    myDrive = pathlib.Path.home().drive
    #mypath = 'e:\\CMUHCH\\print.csv'
    #mypath = myDrive + '\\CMUHCH\\print.csv'
    #myDir = 'e:\\CMUHCH\\'
    #myDir = myDrive + "\\CMUHCH\\" + current_file0
    #myDir = "e:\\CMUHCH\\" + current_file
    #print("myDir", myDir)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
        # 'outputs': myDir,
    })


@excelTable.route("/exportToExcelForStorage", methods=['POST'])
def export_to_Excel_for_Storage():
    print("exportToExcelForStorage....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    #current_file = '..\\領用記錄查詢_'+today + '.xlsx'
    current_file = 'e:\\CMUHCH\\入庫記錄查詢_'+today + '.xlsx'
    #current_file0 = '領用記錄查詢_'+today + '.xlsx'

    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '入庫記錄查詢-' + _name                # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            temp_array = []
            temp_array.append(obj['reqRecord_reagID'])
            temp_array.append(obj['reqRecord_reagName'])
            temp_array.append(obj['reqRecord_supplier'])
            temp_array.append(obj['reqRecord_stockInDate'])
            #temp_array.append(obj['reqRecord_Date'])
            temp_array.append(obj['reqRecord_Employer'])
            # temp_str = obj['reqRecord_cnt'] + obj['reqRecord_unit']
            #temp_array.append(obj['reqRecord_cnt'])
            temp_array.append(obj['reqRecord_ori_count'])
            temp_array.append(obj['reqRecord_In_Unit'])

            ws.append(temp_array)
            # print("obj: ", temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True
            # dim = openpyxl.worksheet.dimensions.ColumnDimension(
            #    ws, index=column, bestFit=True, customWidth=True)
            # ws.column_dimensions[column] = dim

        wb.save(current_file)

    myDrive = pathlib.Path.home().drive
    #mypath = 'e:\\CMUHCH\\print.csv'
    #mypath = myDrive + '\\CMUHCH\\print.csv'
    #myDir = 'e:\\CMUHCH\\'
    #myDir = myDrive + "\\CMUHCH\\" + current_file0
    #myDir = "e:\\CMUHCH\\" + current_file
    #print("myDir", myDir)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
        # 'outputs': myDir,
    })


@excelTable.route("/exportToExcelForStock", methods=['POST'])
def export_to_Excel_for_Stock():
    print("exportToExcelForStock....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    #current_file = '..\\庫存記錄查詢_'+today + '.xlsx'
    current_file = 'e:\\庫存記錄查詢_'+today + '.xlsx'
    #current_file0 = '庫存記錄查詢_'+today + '.xlsx'

    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '庫存記錄查詢-' + _name                # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            temp_array = []

            temp_array.append(obj['stkRecord_reagID'])
            temp_array.append(obj['stkRecord_reagName'])
            temp_array.append(obj['stkRecord_supplier'])
            temp_array.append(obj['stkRecord_Date'])  # 入庫日期
            temp_array.append(obj['stkRecord_period'])  # 效期
            temp_array.append(obj['stkRecord_saftStockUnit'])
            temp_array.append(obj['stkRecord_cntUnit'])
            temp_array.append(obj['stkRecord_grid'])

            ws.append(temp_array)
            # print("obj: ", temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True

        wb.save(current_file)

    myDrive = pathlib.Path.home().drive
    #mypath = 'e:\\CMUHCH\\print.csv'
    #mypath = myDrive + '\\CMUHCH\\print.csv'
    #myDir = 'e:\\CMUHCH\\'
    #myDir = myDrive + "\\CMUHCH\\" + current_file0
    #myDir = "e:\\CMUHCH\\" + current_file0
    #print("myDir", myDir)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
        # 'outputs': myDir,
    })


'''
@excelTable.route("/exportToExcelForInv", methods=['POST'])
def export_to_Excel_for_Inventory():
    print("exportToExcelForInv....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    current_file = '..\\盤點作業_'+today + '.xlsx'
    # print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '庫存記錄查詢-' + _name        # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        s = Session()

        for obj in _blocks:
            temp_array = []
            temp_array.append(obj['id'])
            temp_array.append(obj['stockInTag_reagID'])
            temp_array.append(obj['stockInTag_reagName'])
            temp_array.append(obj['stockInTag_reagPeriod'])  # 效期
            temp_array.append(obj['stockInTag_reagTemp'])   # 保存溫度
            temp_array.append(obj['stockInTag_Date'])       # 入庫日期
            temp_array.append(obj['stockInTag_Employer'])   # 入庫人員

            # temp_array.append(obj['stockInTag_grid'])       # 儲位
            # temp_array.append(obj['stockInTag_cnt'])        # 數量
            # temp_array.append(obj['stockInTag_comment'])    # 說明
            ##
            # ws.append(temp_array)

            gridID = 0
            if obj['intag_id'] != "":
                if obj['isGridChange']:  # 儲位有變更
                    gridID = modify_InTags_grid(obj['stockInTag_grid_id'], int(obj['stockInTag_grid_station']),
                                                int(obj['stockInTag_grid_layout']), int(obj['stockInTag_grid_pos']), int(obj['stockInTag_reagID']))
                    print("return grid id: ", gridID)
                    intag = s.query(InTag).filter_by(
                        id=obj['intag_id']).first()

                    if gridID == -1:
                        return_value = False  # 已經放其他試劑, 儲位重複

                    if gridID > 0:  # 新儲位或空儲位
                        intag.grid_id = gridID
                    intag.count = int(obj['stockInTag_cnt'])      # 修改入(在)庫數量資料
                    intag.comment = obj['stockInTag_comment']     # 修改盤點說明資料
                    intag.updated_at = datetime.datetime.utcnow()  # 資料修改的時間
                    s.commit()

            if gridID >= 0:  # 新儲位或空儲位(>0), 相同儲位(=0)
                temp_array.append(obj['stockInTag_grid'])       # 儲位

            temp_array.append(obj['stockInTag_cnt'])        # 數量
            temp_array.append(obj['stockInTag_comment'])    # 說明

            ws.append(temp_array)

        s.close()

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True

        wb.save(current_file)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
    })
'''


@excelTable.route("/exportToExcelForInv", methods=['POST'])
def export_to_Excel_for_Inventory():
    print("exportToExcelForInv....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    #current_file = '..\\盤點作業_'+today + '.xlsx'
    current_file = 'e:\\盤點作業_'+today + '.xlsx'

    # print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '庫存記錄查詢-' + _name        # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            print("exce obj: ", obj)
            temp_array = []
            temp_array.append(obj['id'])
            temp_array.append(obj['stockInTag_reagID'])
            temp_array.append(obj['stockInTag_reagName'])
            temp_array.append(obj['stockInTag_reagProduct'])  #類別, # 2023-05-23 add
            temp_array.append(obj['stockInTag_stockInBatch']) #批次, # 2023-02-16 add
            temp_array.append(obj['stockInTag_reagPeriod'])  # 效期
            temp_array.append(obj['stockInTag_reagTemp'])   # 保存溫度
            temp_array.append(obj['stockInTag_Date'])       # 入庫日期
            temp_array.append(obj['stockInTag_Employer'])   # 入庫人員

            temp_array.append(obj['stockInTag_grid'])       # 儲位
            temp_array.append(obj['stockInTag_cnt'])        # 在庫數
            temp_array.append(obj['stockInTag_cnt_inv_mdf'])  # 盤點數
            temp_array.append(obj['stockInTag_comment'])    # 說明

            ws.append(temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True

        wb.save(current_file)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
    })
